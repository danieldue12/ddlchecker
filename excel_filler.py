
from openai import OpenAI
from dotenv import load_dotenv
import os
from openpyxl import load_workbook

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

C8_C12 = [
    "¿El paciente tiene antecedentes médicos relevantes?",
    "¿El paciente consume sustancias psicoactivas?",
    "¿El paciente realiza actividad física regularmente?",
    "¿El paciente tiene un entorno familiar estable?",
    "¿El paciente duerme al menos 6 horas por noche?"
]

C14 = ["¿Acepta el tratamiento sugerido?"]
C16_C36 = [f"Condición #{i}" for i in range(1, 22)]
C38_C39 = ["¿Está bajo tratamiento actual?", "¿Presenta síntomas activos?"]

def ask_gpt(question, context):
    prompt = f"""Responde con 1 si la afirmación es verdadera para el paciente, o con 0 si no lo es. No inventes datos.
Contexto: {context[:3000]}
Pregunta: {question}
Respuesta:"""
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[{"role": "user", "content": prompt}],
        temperature=0
    )
    return int(response.choices[0].message.content.strip()[0])

def fill_excel(xlsx_path, context):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    celdas_1 = ["C8", "C9", "C10", "C11", "C12"]
    celdas_2 = ["C14"]
    celdas_3 = [f"C{i}" for i in range(16, 37)]
    celdas_4 = ["C38", "C39"]

    for i, celda in enumerate(celdas_1):
        ws[celda] = ask_gpt(C8_C12[i], context)

    cumplidas1 = sum(ws[celda].value for celda in celdas_1)
    ws["C13"] = int((cumplidas1 / len(celdas_1)) * 100)

    for i, celda in enumerate(celdas_2):
        ws[celda] = ask_gpt(C14[i], context)

    for i, celda in enumerate(celdas_3):
        ws[celda] = ask_gpt(C16_C36[i], context)

    cumplidas3 = sum(ws[celda].value for celda in celdas_3)
    ws["C37"] = int((cumplidas3 / len(celdas_3)) * 100)

    for i, celda in enumerate(celdas_4):
        ws[celda] = ask_gpt(C38_C39[i], context)

    nombre = "paciente"
    for line in context.split("\n"):
        if "Nombre" in line:
            nombre = line.split(":")[-1].strip().replace(" ", "_")
            break

    output_filename = f"lista_chequeo_{nombre}.xlsx"
    wb.save(output_filename)
    return output_filename
